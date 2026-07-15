from __future__ import annotations

import json
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
WORKBOOK = PROJECT_DIR / "Input" / "Codex Manufacturing Data v1.2.xlsx"
EXCLUDED_SHEETS = {"Long Format", "New Batch Entry Template", "Specs"}
LONG_FORMAT_SHEET = "Long Format"
HEADERS = ["Batch", "Day", "Category", "Measurement Type", "Value", "Construct"]


def workbook_unavailable_message() -> str:
    return (
        "The source workbook is open or locked, so the dashboard cannot refresh. "
        "Save and close 'Codex Manufacturing Data v1.2.xlsx', then click Refresh Data again."
    )


def ensure_workbook_writable() -> None:
    try:
        with WORKBOOK.open("r+b"):
            pass
    except OSError:
        sys.exit(workbook_unavailable_message())


def normalize(value: Any) -> str:
    return str(value or "").strip()


def has_value(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def key_for(row: tuple[Any, ...]) -> tuple[str, str]:
    return normalize(row[0]), normalize(row[3])


def backup_workbook() -> Path:
    timestamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup = WORKBOOK.with_name(f"{WORKBOOK.stem}.backup-before-long-format-sync-{timestamp}{WORKBOOK.suffix}")
    shutil.copy2(WORKBOOK, backup)
    return backup


def ensure_long_format_sheet(workbook):
    if LONG_FORMAT_SHEET not in workbook.sheetnames:
        sheet = workbook.create_sheet(LONG_FORMAT_SHEET, 0)
        sheet.append(HEADERS)
        return sheet

    sheet = workbook[LONG_FORMAT_SHEET]
    for index, header in enumerate(HEADERS, start=1):
        if sheet.cell(row=1, column=index).value in (None, ""):
            sheet.cell(row=1, column=index).value = header
    return sheet


def is_batch_sheet(sheet) -> bool:
    headers = [normalize(sheet.cell(row=1, column=column).value).casefold() for column in range(1, 7)]
    return headers == [header.casefold() for header in HEADERS]


def batch_sheet_names(workbook) -> list[str]:
    return [name for name in workbook.sheetnames if name not in EXCLUDED_SHEETS and is_batch_sheet(workbook[name])]


def source_rows_from_batch_sheets(workbook, sheets: list[str]) -> tuple[dict[tuple[str, str], tuple[Any, ...]], dict[str, int], dict[str, int]]:
    rows_by_key: dict[tuple[str, str], tuple[Any, ...]] = {}
    synced_by_sheet = {name: 0 for name in sheets}
    duplicate_skips = {name: 0 for name in sheets}

    for sheet_name in sheets:
        sheet = workbook[sheet_name]
        seen_in_sheet: set[tuple[str, str]] = set()
        for row in sheet.iter_rows(min_row=2, max_col=6, values_only=True):
            trimmed = tuple(row[:6])
            key = key_for(trimmed)
            if not key[0] or not key[1] or not has_value(trimmed[4]):
                continue
            if key in seen_in_sheet or key in rows_by_key:
                duplicate_skips[sheet_name] += 1
                continue
            seen_in_sheet.add(key)
            rows_by_key[key] = trimmed
            synced_by_sheet[sheet_name] += 1

    return rows_by_key, synced_by_sheet, duplicate_skips


def long_format_index(sheet, sync_keys: set[tuple[str, str]]) -> tuple[dict[tuple[str, str], int], list[int]]:
    index: dict[tuple[str, str], int] = {}
    duplicate_rows: list[int] = []

    for row_index in range(2, sheet.max_row + 1):
        row = tuple(sheet.cell(row=row_index, column=column).value for column in range(1, 7))
        key = key_for(row)
        if not key[0] or not key[1]:
            continue
        if key not in sync_keys:
            continue
        if key in index:
            duplicate_rows.append(row_index)
        else:
            index[key] = row_index

    return index, duplicate_rows


def write_row(sheet, row_index: int, row: tuple[Any, ...]) -> None:
    for column, value in enumerate(row[:6], start=1):
        sheet.cell(row=row_index, column=column).value = value


def main() -> None:
    ensure_workbook_writable()
    backup = backup_workbook()
    workbook = load_workbook(WORKBOOK)
    long_format = ensure_long_format_sheet(workbook)
    sheets = batch_sheet_names(workbook)
    rows_by_key, synced_by_sheet, source_duplicate_skips = source_rows_from_batch_sheets(workbook, sheets)
    index, duplicate_long_rows = long_format_index(long_format, set(rows_by_key))

    updated = 0
    added = 0
    for key, row in rows_by_key.items():
        row_index = index.get(key)
        if row_index:
            write_row(long_format, row_index, row)
            updated += 1
        else:
            long_format.append(row)
            index[key] = long_format.max_row
            added += 1

    for row_index in sorted(duplicate_long_rows, reverse=True):
        long_format.delete_rows(row_index, 1)

    try:
        workbook.save(WORKBOOK)
    except OSError:
        sys.exit(workbook_unavailable_message())

    print(
        json.dumps(
            {
                "workbook": str(WORKBOOK),
                "backup": str(backup),
                "batchSheets": sheets,
                "syncedRowsBySheet": synced_by_sheet,
                "updatedRows": updated,
                "addedRows": added,
                "skippedDuplicateSourceRowsBySheet": source_duplicate_skips,
                "removedDuplicateLongFormatRows": len(duplicate_long_rows),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
